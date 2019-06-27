#Trackma AnimeInfoExtractor.py tester v0.9.0.1

# For testing, comment out the last line of the "def __init__(self, filename):" function in AnimeInfoExtractor.py before use (This has already been done in the included sample version)

# Put filenames to be tested in the first column of 'files.xlsx' (The included version is prefilled with random test names, just replace with your own)
# Run anime_extractor_test.py
# The output will be two Excel documents, 'output YYYY.MM.DD  HH.MM.SS.xlsx' and 'summary YYYY.MM.DD  HH.MM.SS.xlsx' where  YYYY.MM.DD  HH.MM.SS are the current time and date:
# The summary spreadsheet shows only the final results provided by the AnimeInfoExtractor to Trackma for each filename
# The output spreadsheet has a sheet for each filename, along with the values of each attribute after each function AnimeInfoExtractor calls is run, allowing you to see precisely where any issue in recognition is occuring.

import AnimeInfoExtractor as AnimeInfoExtractor
import os
from datetime import datetime, date, time
import openpyxl
import re
from decimal import Decimal
from pathlib import Path

path = (os.path.dirname(os.path.abspath(__file__)) + r"\\")


fnlistpath = path + "files.xlsx"
templatepath = path + "output_template.xlsx"

fnfile = openpyxl.load_workbook(fnlistpath)
outputfile = openpyxl.Workbook()
summaryfile = openpyxl.Workbook()
templatefile = openpyxl.load_workbook(templatepath)

fnlist = fnfile.active
outputsheet = outputfile.active
summarysheet = summaryfile.active
templatesheet = templatefile.active

fnlistmax = fnlist.max_row
templatecmax = templatesheet.max_column
templatermax = templatesheet.max_row

for z in range(1, templatecmax + 1):                                            # Initialize Summary Sheet
    celltemplate = templatesheet.cell(row = 1, column = z+1)
    cellcopy = summarysheet.cell(row = z, column = 1)
    cellcopy.value = celltemplate.value

for i in range(1, fnlistmax + 1):                                               # File Sheet Creation Start
    fnamecoord = fnlist.cell(row = i, column = 1)
    fname = fnamecoord.value
    outputfile.create_sheet(index = i , title = "new")
    
    outputfile.active = (i)
    outputsheet = outputfile.active
    
    
    for x in range(1, templatecmax + 1):                                        #Initialize File Sheet
        for y in range(1, templatermax + 1):
            celltemplate = templatesheet.cell(row = y, column = x)
            cellcopy = outputsheet.cell(row = y, column = x)
            cellcopy.value = celltemplate.value
    
    file = fname
    series = AnimeInfoExtractor.AnimeInfoExtractor(fname)
    outputsheet['B2'] = series.originalFilename
    outputsheet['C2'] = series.resolution
    outputsheet['D2'] = series.hash
    outputsheet['E2'] = series.subberTag
    outputsheet['F2'] = ' '.join(series.videoType)
    outputsheet['G2'] = ' '.join(series.audioType)
    outputsheet['H2'] = ' '.join( series.releaseSource)
    outputsheet['I2'] = series.extension
    outputsheet['J2'] = series.episodeStart
    outputsheet['K2'] = series.episodeEnd
    outputsheet['L2'] = series.volumeStart
    outputsheet['M2'] = series.volumeEnd
    outputsheet['N2'] = series.version
    outputsheet['O2'] = series.name
    outputsheet['P2'] = series.pv
    outputsheet['Q2'] = file
    
    file = series.originalFilename
    outputsheet['B3'] = series.originalFilename
    outputsheet['C3'] = series.resolution
    outputsheet['D3'] = series.hash
    outputsheet['E3'] = series.subberTag
    outputsheet['F3'] = ' '.join(series.videoType)
    outputsheet['G3'] = ' '.join(series.audioType)
    outputsheet['H3'] = ' '.join( series.releaseSource)
    outputsheet['I3'] = series.extension
    outputsheet['J3'] = series.episodeStart
    outputsheet['K3'] = series.episodeEnd
    outputsheet['L3'] = series.volumeStart
    outputsheet['M3'] = series.volumeEnd
    outputsheet['N3'] = series.version
    outputsheet['O3'] = series.name
    outputsheet['P3'] = series.pv
    outputsheet['Q3'] = file
    
    file = series._AnimeInfoExtractor__testIfEp(file)               # Test if file is an episode or Special/Supplementary File, and force a failure condiion if this is the case, to avoid false positives
    outputsheet['B4'] = series.originalFilename
    outputsheet['C4'] = series.resolution
    outputsheet['D4'] = series.hash
    outputsheet['E4'] = series.subberTag
    outputsheet['F4'] = ' '.join(series.videoType)
    outputsheet['G4'] = ' '.join(series.audioType)
    outputsheet['H4'] = ' '.join( series.releaseSource)
    outputsheet['I4'] = series.extension
    outputsheet['J4'] = series.episodeStart
    outputsheet['K4'] = series.episodeEnd
    outputsheet['L4'] = series.volumeStart
    outputsheet['M4'] = series.volumeEnd
    outputsheet['N4'] = series.version
    outputsheet['O4'] = series.name
    outputsheet['P4'] = series.pv
    outputsheet['Q4'] = file
    
    file = series._AnimeInfoExtractor__extractExtension(file)
    outputsheet['B5'] = series.originalFilename
    outputsheet['C5'] = series.resolution
    outputsheet['D5'] = series.hash
    outputsheet['E5'] = series.subberTag
    outputsheet['F5'] = ' '.join(series.videoType)
    outputsheet['G5'] = ' '.join(series.audioType)
    outputsheet['H5'] = ' '.join( series.releaseSource)
    outputsheet['I5'] = series.extension
    outputsheet['J5'] = series.episodeStart
    outputsheet['K5'] = series.episodeEnd
    outputsheet['L5'] = series.volumeStart
    outputsheet['M5'] = series.volumeEnd
    outputsheet['N5'] = series.version
    outputsheet['O5'] = series.name
    outputsheet['P5'] = series.pv
    outputsheet['Q5'] = file
    
    file = series._AnimeInfoExtractor__cleanUpSpaces(file)
    outputsheet['B6'] = series.originalFilename
    outputsheet['C6'] = series.resolution
    outputsheet['D6'] = series.hash
    outputsheet['E6'] = series.subberTag
    outputsheet['F6'] = ' '.join(series.videoType)
    outputsheet['G6'] = ' '.join(series.audioType)
    outputsheet['H6'] = ' '.join( series.releaseSource)
    outputsheet['I6'] = series.extension
    outputsheet['J6'] = series.episodeStart
    outputsheet['K6'] = series.episodeEnd
    outputsheet['L6'] = series.volumeStart
    outputsheet['M6'] = series.volumeEnd
    outputsheet['N6'] = series.version
    outputsheet['O6'] = series.name
    outputsheet['P6'] = series.pv
    outputsheet['Q6'] = file
    
    file = series._AnimeInfoExtractor__extractSpecialTags(file)
    outputsheet['B7'] = series.originalFilename
    outputsheet['C7'] = series.resolution
    outputsheet['D7'] = series.hash
    outputsheet['E7'] = series.subberTag
    outputsheet['F7'] = ' '.join(series.videoType)
    outputsheet['G7'] = ' '.join(series.audioType)
    outputsheet['H7'] = ' '.join( series.releaseSource)
    outputsheet['I7'] = series.extension
    outputsheet['J7'] = series.episodeStart
    outputsheet['K7'] = series.episodeEnd
    outputsheet['L7'] = series.volumeStart
    outputsheet['M7'] = series.volumeEnd
    outputsheet['N7'] = series.version
    outputsheet['O7'] = series.name
    outputsheet['P7'] = series.pv
    outputsheet['Q7'] = file
        
    file = series._AnimeInfoExtractor__extractVideoProfile(file)
    outputsheet['B8'] = series.originalFilename
    outputsheet['C8'] = series.resolution
    outputsheet['D8'] = series.hash
    outputsheet['E8'] = series.subberTag
    outputsheet['F8'] = ' '.join(series.videoType)
    outputsheet['G8'] = ' '.join(series.audioType)
    outputsheet['H8'] = ' '.join( series.releaseSource)
    outputsheet['I8'] = series.extension
    outputsheet['J8'] = series.episodeStart
    outputsheet['K8'] = series.episodeEnd
    outputsheet['L8'] = series.volumeStart
    outputsheet['M8'] = series.volumeEnd
    outputsheet['N8'] = series.version
    outputsheet['O8'] = series.name
    outputsheet['P8'] = series.pv
    outputsheet['Q8'] = file
    
    file = series._AnimeInfoExtractor__extractResolution(file)
    outputsheet['B9'] = series.originalFilename
    outputsheet['C9'] = series.resolution
    outputsheet['D9'] = series.hash
    outputsheet['E9'] = series.subberTag
    outputsheet['F9'] = ' '.join(series.videoType)
    outputsheet['G9'] = ' '.join(series.audioType)
    outputsheet['H9'] = ' '.join( series.releaseSource)
    outputsheet['I9'] = series.extension
    outputsheet['J9'] = series.episodeStart
    outputsheet['K9'] = series.episodeEnd
    outputsheet['L9'] = series.volumeStart
    outputsheet['M9'] = series.volumeEnd
    outputsheet['N9'] = series.version
    outputsheet['O9'] = series.name
    outputsheet['P9'] = series.pv
    outputsheet['Q9'] = file
    
    file = series._AnimeInfoExtractor__extractHash(file)
    outputsheet['B10'] = series.originalFilename
    outputsheet['C10'] = series.resolution
    outputsheet['D10'] = series.hash
    outputsheet['E10'] = series.subberTag
    outputsheet['F10'] = ' '.join(series.videoType)
    outputsheet['G10'] = ' '.join(series.audioType)
    outputsheet['H10'] = ' '.join( series.releaseSource)
    outputsheet['I10'] = series.extension
    outputsheet['J10'] = series.episodeStart
    outputsheet['K10'] = series.episodeEnd
    outputsheet['L10'] = series.volumeStart
    outputsheet['M10'] = series.volumeEnd
    outputsheet['N10'] = series.version
    outputsheet['O10'] = series.name
    outputsheet['P10'] = series.pv
    outputsheet['Q10'] = file    
    
    remux = series._AnimeInfoExtractor__checkIfRemux(file)
    outputsheet['B11'] = series.originalFilename
    outputsheet['C11'] = series.resolution
    outputsheet['D11'] = series.hash
    outputsheet['E11'] = series.subberTag
    outputsheet['F11'] = ' '.join(series.videoType)
    outputsheet['G11'] = ' '.join(series.audioType)
    outputsheet['H11'] = ' '.join( series.releaseSource)
    outputsheet['I11'] = series.extension
    outputsheet['J11'] = series.episodeStart
    outputsheet['K11'] = series.episodeEnd
    outputsheet['L11'] = series.volumeStart
    outputsheet['M11'] = series.volumeEnd
    outputsheet['N11'] = series.version
    outputsheet['O11'] = series.name
    outputsheet['P11'] = series.pv
    outputsheet['Q11'] = file
    
    file = series._AnimeInfoExtractor__cleanUpBrackets(file)
    outputsheet['B12'] = series.originalFilename
    outputsheet['C12'] = series.resolution
    outputsheet['D12'] = series.hash
    outputsheet['E12'] = series.subberTag
    outputsheet['F12'] = ' '.join(series.videoType)
    outputsheet['G12'] = ' '.join(series.audioType)
    outputsheet['H12'] = ' '.join( series.releaseSource)
    outputsheet['I12'] = series.extension
    outputsheet['J12'] = series.episodeStart
    outputsheet['K12'] = series.episodeEnd
    outputsheet['L12'] = series.volumeStart
    outputsheet['M12'] = series.volumeEnd
    outputsheet['N12'] = series.version
    outputsheet['O12'] = series.name
    outputsheet['P12'] = series.pv
    outputsheet['Q12'] = file
    
    file = series._AnimeInfoExtractor__extractSubber(file, remux)
    outputsheet['B13'] = series.originalFilename
    outputsheet['C13'] = series.resolution
    outputsheet['D13'] = series.hash
    outputsheet['E13'] = series.subberTag
    outputsheet['F13'] = ' '.join(series.videoType)
    outputsheet['G13'] = ' '.join(series.audioType)
    outputsheet['H13'] = ' '.join( series.releaseSource)
    outputsheet['I13'] = series.extension
    outputsheet['J13'] = series.episodeStart
    outputsheet['K13'] = series.episodeEnd
    outputsheet['L13'] = series.volumeStart
    outputsheet['M13'] = series.volumeEnd
    outputsheet['N13'] = series.version
    outputsheet['O13'] = series.name
    outputsheet['P13'] = series.pv
    outputsheet['Q13'] = file
    
    file = series._AnimeInfoExtractor__extractVersion(file)
    outputsheet['B14'] = series.originalFilename
    outputsheet['C14'] = series.resolution
    outputsheet['D14'] = series.hash
    outputsheet['E14'] = series.subberTag
    outputsheet['F14'] = ' '.join(series.videoType)
    outputsheet['G14'] = ' '.join(series.audioType)
    outputsheet['H14'] = ' '.join( series.releaseSource)
    outputsheet['I14'] = series.extension
    outputsheet['J14'] = series.episodeStart
    outputsheet['K14'] = series.episodeEnd
    outputsheet['L14'] = series.volumeStart
    outputsheet['M14'] = series.volumeEnd
    outputsheet['N14'] = series.version
    outputsheet['O14'] = series.name
    outputsheet['P14'] = series.pv
    outputsheet['Q14'] = file
    
    # Store the possible length of the title
    title_len = len(file)
    outputsheet['B15'] = series.originalFilename
    outputsheet['C15'] = series.resolution
    outputsheet['D15'] = series.hash
    outputsheet['E15'] = series.subberTag
    outputsheet['F15'] = ' '.join(series.videoType)
    outputsheet['G15'] = ' '.join(series.audioType)
    outputsheet['H15'] = ' '.join( series.releaseSource)
    outputsheet['I15'] = series.extension
    outputsheet['J15'] = series.episodeStart
    outputsheet['K15'] = series.episodeEnd
    outputsheet['L15'] = series.volumeStart
    outputsheet['M15'] = series.volumeEnd
    outputsheet['N15'] = series.version
    outputsheet['O15'] = series.name
    outputsheet['P15'] = series.pv
    outputsheet['Q15'] = file
    
    file, title_len = series._AnimeInfoExtractor__extractVolumeIfPack(file, title_len)
    outputsheet['B16'] = series.originalFilename
    outputsheet['C16'] = series.resolution
    outputsheet['D16'] = series.hash
    outputsheet['E16'] = series.subberTag
    outputsheet['F16'] = ' '.join(series.videoType)
    outputsheet['G16'] = ' '.join(series.audioType)
    outputsheet['H16'] = ' '.join( series.releaseSource)
    outputsheet['I16'] = series.extension
    outputsheet['J16'] = series.episodeStart
    outputsheet['K16'] = series.episodeEnd
    outputsheet['L16'] = series.volumeStart
    outputsheet['M16'] = series.volumeEnd
    outputsheet['N16'] = series.version
    outputsheet['O16'] = series.name
    outputsheet['P16'] = series.pv
    outputsheet['Q16'] = file
    
    file = series._AnimeInfoExtractor__extractPv(file)
    outputsheet['B17'] = series.originalFilename
    outputsheet['C17'] = series.resolution
    outputsheet['D17'] = series.hash
    outputsheet['E17'] = series.subberTag
    outputsheet['F17'] = ' '.join(series.videoType)
    outputsheet['G17'] = ' '.join(series.audioType)
    outputsheet['H17'] = ' '.join( series.releaseSource)
    outputsheet['I17'] = series.extension
    outputsheet['J17'] = series.episodeStart
    outputsheet['K17'] = series.episodeEnd
    outputsheet['L17'] = series.volumeStart
    outputsheet['M17'] = series.volumeEnd
    outputsheet['N17'] = series.version
    outputsheet['O17'] = series.name
    outputsheet['P17'] = series.pv
    outputsheet['Q17'] = file
    
    if series.pv == -1:
        file = series._AnimeInfoExtractor__extractEpisodeNumbers(file)
    outputsheet['B18'] = series.originalFilename
    outputsheet['C18'] = series.resolution
    outputsheet['D18'] = series.hash
    outputsheet['E18'] = series.subberTag
    outputsheet['F18'] = ' '.join(series.videoType)
    outputsheet['G18'] = ' '.join(series.audioType)
    outputsheet['H18'] = ' '.join( series.releaseSource)
    outputsheet['I18'] = series.extension
    outputsheet['J18'] = series.episodeStart
    outputsheet['K18'] = series.episodeEnd
    outputsheet['L18'] = series.volumeStart
    outputsheet['M18'] = series.volumeEnd
    outputsheet['N18'] = series.version
    outputsheet['O18'] = series.name
    outputsheet['P18'] = series.pv
    outputsheet['Q18'] = file
    
    # Truncate remainder to title length if needed (for where volume was found)
    file = file[:title_len]
    outputsheet['B19'] = series.originalFilename
    outputsheet['C19'] = series.resolution
    outputsheet['D19'] = series.hash
    outputsheet['E19'] = series.subberTag
    outputsheet['F19'] = ' '.join(series.videoType)
    outputsheet['G19'] = ' '.join(series.audioType)
    outputsheet['H19'] = ' '.join( series.releaseSource)
    outputsheet['I19'] = series.extension
    outputsheet['J19'] = series.episodeStart
    outputsheet['K19'] = series.episodeEnd
    outputsheet['L19'] = series.volumeStart
    outputsheet['M19'] = series.volumeEnd
    outputsheet['N19'] = series.version
    outputsheet['O19'] = series.name
    outputsheet['P19'] = series.pv
    outputsheet['Q19'] = file
    
    # Strip any trailing opening brackets
    file = file.rstrip('([{')
    outputsheet['B20'] = series.originalFilename
    outputsheet['C20'] = series.resolution
    outputsheet['D20'] = series.hash
    outputsheet['E20'] = series.subberTag
    outputsheet['F20'] = ' '.join(series.videoType)
    outputsheet['G20'] = ' '.join(series.audioType)
    outputsheet['H20'] = ' '.join( series.releaseSource)
    outputsheet['I20'] = series.extension
    outputsheet['J20'] = series.episodeStart
    outputsheet['K20'] = series.episodeEnd
    outputsheet['L20'] = series.volumeStart
    outputsheet['M20'] = series.volumeEnd
    outputsheet['N20'] = series.version
    outputsheet['O20'] = series.name
    outputsheet['P20'] = series.pv
    outputsheet['Q20'] = file
    
    series._AnimeInfoExtractor__extractShowName(file)
    outputsheet['B21'] = series.originalFilename
    outputsheet['C21'] = series.resolution
    outputsheet['D21'] = series.hash
    outputsheet['E21'] = series.subberTag
    outputsheet['F21'] = ' '.join(series.videoType)
    outputsheet['G21'] = ' '.join(series.audioType)
    outputsheet['H21'] = ' '.join( series.releaseSource)
    outputsheet['I21'] = series.extension
    outputsheet['J21'] = series.episodeStart
    outputsheet['K21'] = series.episodeEnd
    outputsheet['L21'] = series.volumeStart
    outputsheet['M21'] = series.volumeEnd
    outputsheet['N21'] = series.version
    outputsheet['O21'] = series.name
    outputsheet['P21'] = series.pv
    outputsheet['Q21'] = file
    

    for a in range(1, 17):  
        cellextract = outputsheet.cell(row = 21, column = a+1)
        cellcopy = summarysheet.cell(row = a, column = i+1)
        cellcopy.value = cellextract.value
    
    summarysheet.title = "Summary"
    outputsheet.title = file[:20]

d = datetime.now()                                      # Get current date and time
dt = d.strftime(" %Y.%m.%d-%H.%M.%S")                   # Create date/time string
summarypath = (path + "summary" +dt + ".xlsx")          # Set path and filename for Summary file
outputpath = (path + "output" +dt + ".xlsx")            # Set path and filename for Output file
summaryfile.save(summarypath)                           # Save Summary file
outputfile.save(outputpath)                             # Save Output file